"""Load control libraries from the AI Governance Playbook workbook and write results back."""
from __future__ import annotations

import datetime as dt
import warnings
from dataclasses import dataclass, asdict

from openpyxl import load_workbook

warnings.filterwarnings("ignore", message="Data Validation extension")

# sheet name -> (id column, title column, requirement column, owner column, crosswalk columns)
LIBRARIES = {
    "MAS": ("Control Library - MAS", "MAS ref", "MAS expectation area", "Primary MAS source", "Owner (role)", ["ISO 42001", "NIST AI RMF"]),
    "MGF Agentic": ("Control Library - MGF Agentic", "MGF Control ID", "Control / Recommended Measure", "Objective — what it bounds", "Owner (role)", ["MAS AIRG theme", "ISO/IEC 42001 Annex A"]),
    "SAFR": ("Control Library - SAFR", "SAFR Control ID", "Control / recommended measure", "Objective — what it bounds", "Owner (role)", ["MAS AIRG theme", "ISO/IEC 42001 Annex A"]),
    "ISO 42001": ("Control Library - ISO 42001", "Control ID", "Control Title", "Requirement (summary)", "Owner (role)", []),
    "NIST AI RMF": ("Control Library - NIST AI RMF", "Subcategory", "Category", "Subcategory outcome", "Owner (role)", []),
}
WRITEBACK_COLS = ["Status", "Maturity (1-5)", "Evidence / artifact", "Last reviewed", "Gap / notes"]


@dataclass
class Control:
    key: str
    lib: str
    id: str
    title: str
    req: str
    owner: str
    maps: str
    row: int  # 1-based worksheet row, used for write-back

    def to_dict(self):
        return asdict(self)


def _header_map(ws, header_row=3):
    return {str(c.value).strip(): c.column for c in ws[header_row] if c.value}


def load_controls(path_or_file) -> dict[str, list[Control]]:
    wb = load_workbook(path_or_file, read_only=False, data_only=True)
    out: dict[str, list[Control]] = {}
    for lib, (sheet, id_col, title_col, req_col, owner_col, xw) in LIBRARIES.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        h = _header_map(ws)
        if id_col not in h:
            continue
        rows = []
        for r in range(4, ws.max_row + 1):
            cid = ws.cell(r, h[id_col]).value
            if not cid:
                continue
            get = lambda name: str(ws.cell(r, h[name]).value or "").strip() if name in h else ""
            maps = " | ".join(f"{c}: {get(c)}" for c in xw if get(c))
            req = get(req_col)
            if lib == "MAS":
                req = f"Expectation: {get(title_col)}. Source: {req}"
            rows.append(Control(f"{lib}::{str(cid).strip()}", lib, str(cid).strip(), get(title_col), req, get(owner_col), maps, r))
        out[lib] = rows
    return out


def write_back(src_path_or_file, dst_path: str, controls: dict[str, list[Control]], decisions: dict, ai: dict, evidence: dict) -> int:
    """Write accepted decisions into a copy of the workbook. Returns number of rows updated."""
    wb = load_workbook(src_path_or_file)
    n = 0
    for lib, rows in controls.items():
        ws = wb[LIBRARIES[lib][0]]
        h = _header_map(ws)
        for c in rows:
            d = decisions.get(c.key)
            if not d:
                continue
            a = ai.get(c.key, {})
            ev = evidence.get(c.key, {})
            vals = {
                "Status": {"full": "Evidenced", "partial": "Partially evidenced", "none": "Not evidenced"}[d["sufficiency"]],
                "Maturity (1-5)": d["maturity"],
                "Evidence / artifact": (ev.get("file_name") or "") + (" — " if ev.get("file_name") and ev.get("text") else "") + (ev.get("text") or "")[:500],
                "Last reviewed": dt.date.fromisoformat(d["at"][:10]),
                "Gap / notes": "; ".join(a.get("gaps", [])) + (f" | Reviewer: {d['note']}" if d.get("note") else "") + f" | Accepted by {d['reviewer']}",
            }
            for col, v in vals.items():
                if col in h:
                    ws.cell(c.row, h[col]).value = v
            n += 1
    wb.save(dst_path)
    return n
