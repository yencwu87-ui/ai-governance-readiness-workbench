"""
Adapters for lifecycle Plays 10 and 11.

stress_tracker  — reads the playbook's 'Stress-Test Tracker' sheet (xlsx) or a CSV export of it.
                  Records: n, hotspot, class, maps_to, owner, cadence, last_run, result, residual_risk,
                  next_run, max_days (derived from cadence).
decisions_log   — reads the workbench's data/assessments.json (Lane A recorded decisions).
                  Records: key, reviewer, at, sufficiency, maturity, ai_sufficiency, override (bool).
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from . import AdapterResult, _hash_file, _read_table, _resolve, adapter

CADENCE_DAYS = {"quarterly": 100, "semi-annual": 190, "semiannual": 190, "annual": 380, "monthly": 35, "weekly": 10}


def _max_days(cadence: str) -> int | None:
    c = (cadence or "").lower()
    for k, v in CADENCE_DAYS.items():
        if k in c:
            return v
    return None


@adapter("stress_tracker")
def stress_tracker(target: Path, cfg: dict) -> AdapterResult:
    """cfg: {path: xlsx or csv, sheet: 'Stress-Test Tracker'}"""
    p = _resolve(target, cfg["path"])
    if not p.exists():
        # allow a glob for the git-ignored workbook, e.g. data/*.xlsx
        matches = sorted(target.glob(cfg["path"])) if any(ch in cfg["path"] for ch in "*?") else []
        if not matches:
            return AdapterResult("missing", message=f"{cfg['path']} not found")
        p = matches[0]
    rows: list[dict]
    if p.suffix.lower() == ".csv":
        rows = _read_table(p)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(str(p), data_only=True, read_only=True)
        sheet = cfg.get("sheet", "Stress-Test Tracker")
        if sheet not in wb.sheetnames:
            return AdapterResult("missing", message=f"sheet {sheet!r} not in {p.name}")
        raw = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        hdr_i = next((i for i, r in enumerate(raw) if r and str(r[0]).strip() == "#"), None)
        if hdr_i is None:
            return AdapterResult("error", message="tracker header row not found")
        hdr = [str(h).strip() if h is not None else "" for h in raw[hdr_i]]
        rows = [dict(zip(hdr, r)) for r in raw[hdr_i + 1:] if r and r[0] is not None and str(r[0]).strip().isdigit()]
    records = []
    for r in rows:
        g = lambda *names: next((r[n] for n in names if n in r and r[n] not in (None, "")), None)
        cadence = str(g("Cadence", "cadence") or "")
        records.append({
            "n": g("#", "n"), "hotspot": g("Hotspot", "hotspot"), "class": g("Class", "class"),
            "maps_to": g("Maps to", "maps_to"), "owner": g("Owner (role)", "owner"), "cadence": cadence,
            "last_run": g("Last run", "last_run"), "result": str(g("Result", "result") or "").strip(),
            "residual_risk": str(g("Residual risk", "residual_risk") or "").strip(), "next_run": g("Next run", "next_run"),
            "max_days": _max_days(cadence),
        })
    return AdapterResult("ok", records, [_hash_file(p)])


@adapter("decisions_log")
def decisions_log(target: Path, cfg: dict) -> AdapterResult:
    """cfg: {path: data/assessments.json}"""
    p = _resolve(target, cfg.get("path", "data/assessments.json"))
    if not p.exists():
        return AdapterResult("missing", message=f"{p} not found")
    try:
        state = json.loads(p.read_text())
    except json.JSONDecodeError as e:
        return AdapterResult("error", message=str(e))
    records = []
    for key, d in (state.get("decisions") or {}).items():
        ai = d.get("aiSufficiency")
        records.append({
            "key": key, "reviewer": d.get("reviewer"), "at": d.get("at"), "sufficiency": d.get("sufficiency"),
            "maturity": d.get("maturity"), "ai_sufficiency": ai, "ai_maturity": d.get("aiMaturity"),
            "override": bool(ai) and (d.get("sufficiency") != ai or d.get("maturity") != d.get("aiMaturity")),
            "unnamed": (d.get("reviewer") or "").strip().lower() in ("", "unnamed reviewer"),
        })
    return AdapterResult("ok", records, [_hash_file(p)])
