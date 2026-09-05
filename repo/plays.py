"""Read the 'Playbooks & Runbooks' sheet of the playbook workbook into plays and steps.

A play header row carries the trigger/frequency text and a 'Controls satisfied' string like
    'ISO A.6.2.5, A.9.2-A.9.4 | NIST MANAGE 1.1, MEASURE 2.3 | MAS M3.4, M3.9 | MGF D2.3, D2.4, D3.10'
which is parsed (ranges expanded) into {lib: [control ids]}.

Step rows are numbered within the play; a step id is 'P<play>.<step>', e.g. 'P4.2'.
Lane B controls point at these via `play_refs`.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

SHEET = "Playbooks & Runbooks"
LIB_PREFIXES = ("ISO", "NIST", "MAS", "MGF", "FEAT")


@dataclass
class Step:
    id: str          # P4.2
    play: int
    n: int
    owner: str
    action: str
    cadence: str
    evidence: str


@dataclass
class Play:
    id: str          # P4
    n: int
    title: str
    header: str      # trigger / frequency / accountable owner text
    controls: dict[str, list[str]] = field(default_factory=dict)   # lib -> ids
    steps: list[Step] = field(default_factory=list)

    @property
    def control_ids(self) -> set[str]:
        return {c for ids in self.controls.values() for c in ids}


# ---------- control string parsing ----------

_TOKEN = re.compile(r"[A-Z]+(?:\s+[A-Z]+)?\s*[A-Z]?\d+(?:\.\d+)*")


def _expand(a: str, b: str) -> list[str]:
    """Expand 'A.5.2'-'A.5.5', 'D3.1'-'D3.9', '2.1'-'2.13', 'F1'-'F4' when only the last number differs."""
    ma, mb = re.match(r"^(.*?)(\d+)$", a), re.match(r"^(.*?)(\d+)$", b)
    if not (ma and mb) or ma.group(1) != mb.group(1):
        return [a, b]
    lo, hi = int(ma.group(2)), int(mb.group(2))
    if hi < lo or hi - lo > 40:
        return [a, b]
    return [f"{ma.group(1)}{i}" for i in range(lo, hi + 1)]


def parse_controls(s: str) -> dict[str, list[str]]:
    out: dict[str, list[str]] = {}
    for seg in (s or "").split("|"):
        seg = seg.strip()
        lib = next((p for p in LIB_PREFIXES if seg.upper().startswith(p)), None)
        if not lib:
            continue
        body = seg[len(lib):]
        body = re.sub(r"\([^)]*\)", "", body)           # drop '(assess)', '(sign-off)'
        ids: list[str] = []
        family = ""                                       # NIST 'MEASURE', 'GOVERN' carry over commas
        for part in [p.strip() for p in body.split(",") if p.strip()]:
            m = re.match(r"^([A-Z]+)\s+(.*)$", part)      # 'MEASURE 2.1-2.13'
            if m and lib == "NIST":
                family, part = m.group(1), m.group(2)
            if "-" in part and lib != "NIST" or (lib == "NIST" and "-" in part):
                a, b = [x.strip() for x in part.split("-", 1)]
                items = _expand(a, b)
            else:
                items = [part]
            for it in items:
                it = it.strip()
                if not it:
                    continue
                ids.append(f"{family} {it}".strip() if lib == "NIST" and family and not it.startswith(family) else it)
        if ids:
            out[lib] = ids
    return out


# ---------- sheet reading ----------

def load_plays(src) -> list[Play]:
    """src: path or file-like (as passed to playbook.load_controls)."""
    from openpyxl import load_workbook
    wb = load_workbook(src, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        return []
    plays: list[Play] = []
    cur: Play | None = None
    for row in wb[SHEET].iter_rows(values_only=True):
        row = list(row) + [None] * (6 - len(row))
        a = str(row[0]).strip() if row[0] is not None else ""
        m = re.match(r"^PLAY\s+(\d+)\s*[—–-]\s*(.+)$", a)
        if m:
            cur = Play(id=f"P{m.group(1)}", n=int(m.group(1)), title=m.group(2).strip(),
                       header=str(row[1] or "").strip(), controls=parse_controls(str(row[5] or "")))
            plays.append(cur)
            continue
        ms = re.match(r"^Step\s+(\d+)$", a)
        if ms and cur is not None:
            n = int(ms.group(1))
            cur.steps.append(Step(id=f"{cur.id}.{n}", play=cur.n, n=n, owner=str(row[1] or "").strip(),
                                  action=str(row[2] or "").strip(), cadence=str(row[3] or "").strip(),
                                  evidence=str(row[4] or "").strip()))
    return plays


def step_index(plays: list[Play]) -> dict[str, Step]:
    return {s.id: s for p in plays for s in p.steps}


def norm_id(cid: str) -> str:
    """Playbook control id as stored in the libraries -> comparable token ('M3.4 ★' -> 'M3.4')."""
    return str(cid).split()[0].rstrip("★").strip() if cid else ""
